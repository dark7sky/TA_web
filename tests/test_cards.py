import unittest

from openpyxl import Workbook

import cards


class ShinhanLayoutTests(unittest.TestCase):
    def test_uses_amount_column_when_shinhan_sheet_has_new_layout(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "신한2605"
        ws["A1"] = "거래일"
        ws["A2"] = "2026.05.30 10:21"
        ws["B1"] = "카드구분"
        ws["C1"] = "이용카드"
        ws["D1"] = "가맹점명"
        ws["E1"] = "업종"
        ws["F1"] = "승인번호"
        ws["G1"] = "금액"
        ws["H1"] = "매입구분"
        ws["I1"] = "이용구분"
        ws["A2"] = "2026.05.30 10:21"
        ws["B2"] = "신용"
        ws["C2"] = "본인690*"
        ws["D2"] = "네이버플러스 멤버십"
        ws["F2"] = "30823078"
        ws["G2"] = 11500
        ws["H2"] = "결제확정"
        ws["I2"] = "일시불"

        data = cards.card_analysis("신한2605", ws)

        transaction_rows = [row for row in data if row[1] < 0]
        summary_rows = [row for row in data if row[1] > 0]

        self.assertEqual(transaction_rows[0][1], -11500)
        self.assertEqual(summary_rows[-1][1], 11500)
        self.assertEqual(sum(row[1] for row in data), 0)


if __name__ == "__main__":
    unittest.main()
