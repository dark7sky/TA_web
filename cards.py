"""cards.py ??Parse credit card Excel statements and store them in PostgreSQL.

Optimized version (2026-03-04)
Features:
- Parameterized SQL (no injection risk)
- O(N) deduplication (fixed O(N^2) list.remove)
- Removed bare excepts and tightened typing
"""
import copy
import datetime
import os
import pickle
import re
import psycopg2
import os
from dotenv import load_dotenv
load_dotenv()
from typing import Any, List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from logger import Logger

debug_print = False
supportCardVendors = {
    "?좏븳": "shinhan",
    "?쇱꽦": "samsung",
    "?곕━": "woori",
    "KB":   "KB",
    "?섎굹": "hana"
}

logs = Logger(os.path.basename(__file__).split(".")[0])


def log(msg: str) -> None:
    if debug_print:
        print(msg)
    else:
        logs.msg(msg)


def Analysis_row(vendor: str, year: int, day_column: int, row: tuple) -> bool:
    """Validate if the Excel row contains a valid transaction entry."""
    val = row[day_column].value
    if not isinstance(val, str):
        return False
        
    if vendor == "woori":
        if not re.match(r"\d\d\.\d\d \d\d:\d\d:\d\d", val):
            return False
    else:
        if not val.startswith(str(year)):
            return False
        if vendor == "shinhan" and "?댁슜湲덉븸?좎씤" in str(row[2].value):
            return False
            
    return True


def parse_number(val: Any) -> int:
    """Safely parse Excel comma-separated string to integer."""
    if val is None:
        return 0
    return int(str(val).replace(",", ""))


def Analysis_sheet(
    ws: Worksheet,
    vendor: str,
    year: int,
    day_column: int,
    value_column: int,
    vendor_type: str,
    time_format: str,
    fn_tick: datetime.datetime,
) -> List[list]:
    """Parse transaction rows into (datetime, negative_value) lists.
    
    A final summation row is appended at ``fn_tick`` containing the
    sheet total.
    """
    totals = 0
    datas: List[list] = []
    
    for row in ws.rows:
        if vendor == "shinhan" and vendor_type == "1":
            desc = row[2].value
            if not isinstance(desc, str):
                continue
            if "연회비" in desc:
                amt = parse_number(row[value_column].value)
                datas.append([fn_tick - datetime.timedelta(seconds=1), -amt])
                totals += amt
                continue

        if Analysis_row(vendor, year, day_column, row):
            
            # --- Extract transaction value depending on vendor_type ---
            if vendor_type == "1":
                value = parse_number(row[value_column].value) + parse_number(row[value_column + 1].value)
            
            elif vendor_type == "2":
                value = parse_number(row[value_column].value)
                if len(row) > value_column + 2:
                    if row[value_column + 2].value == "?뱀씤痍⑥냼":
                        value = -value
                        
            elif vendor_type == "3":
                value = parse_number(row[value_column].value) - parse_number(row[value_column + 1].value)
                
            elif vendor_type == "4":
                value = parse_number(row[value_column].value) - parse_number(row[value_column + 6].value)
                
            elif vendor_type == "5":
                # KB new format
                value = parse_number(row[value_column].value) - parse_number(row[value_column + 4].value)
                
            elif vendor_type == "6":
                # KB new format 2
                value = parse_number(row[value_column].value) - parse_number(row[value_column + 2].value)
                
            else:
                continue

            # --- Parse transaction date ---
            rday = row[day_column].value
            try:
                if vendor_type == "6" and vendor == "KB":
                    tick = datetime.datetime.strptime(str(rday).split()[0], time_format).replace(hour=23, minute=59, second=59)
                elif vendor == "woori" and vendor_type == "3":
                    tick = datetime.datetime.strptime(f"{year}.{rday}", "%Y.%m.%d %H:%M:%S").replace(hour=23, minute=59, second=59)
                else:
                    tick = datetime.datetime.strptime(str(rday), time_format).replace(hour=23, minute=59, second=59)
            except ValueError as e:
                logs.warning(f"Failed to parse date '{rday}' with format '{time_format}': {e}")
                continue

            # Nudge transaction time by seconds so consecutive days from different vendors don't completely overlap
            if vendor == "KB":
                tick = tick.replace(second=56)
            elif vendor == "woori":
                tick = tick.replace(second=57)
            elif vendor == "samsung":
                tick = tick.replace(second=58)

            totals += value
            
            # If tick matches previous element exactly, aggregate, otherwise append
            if datas and datas[-1][0] == tick:
                datas[-1][1] -= value
            else:
                datas.append([tick, -value])
                
    # Add summary total for the sheet at fn_tick
    datas.append([fn_tick, totals])
    return datas


def card_analysis(sheetname: str, ws: Worksheet) -> list | tuple[str, Exception]:
    """Drive analysis constraints (columns, time formats) per vendor."""
    print(f", {sheetname}", end="", flush=True)
    
    try:
        prefix = sheetname[0:2]
        vendor = supportCardVendors[prefix]
        year = 2000 + int(sheetname[-4:-2])
    except KeyError as e:
        return "ERROR", Exception(f"吏?먮릺吏 ?딅뒗 移대뱶 ?묐몢?? {prefix}")
    except ValueError as e:
        return "ERROR", Exception(f"Sheet ?대쫫 ?곕룄 ?뚯떛 ?먮윭({sheetname}): {e}")

    day_column = 0
    value_column = 0
    vendor_type = ""
    time_format = ""
    fn_tick = datetime.datetime.now()

    if vendor == "shinhan":
        first_id = str(ws["A1"].value)
        if "?댁슜?쇱옄蹂?移대뱶?ъ슜?댁뿭" in first_id:
            # ?붾퀎 ?湲?紐낆꽭??
            day_column = 0
            value_column = 6
            vendor_type = "1"
            time_format = "%Y.%m.%d"
            try:
                fn_tick = datetime.datetime.strptime(str(ws["C4"].value), time_format)
                fn_tick = fn_tick.replace(day=14, hour=17, minute=42, second=0)
            except Exception as e:
                logs.warning(f"Shinhan fn_tick fallback: {e}")
                
        elif "?댁슜?쇱떆" in first_id:
            # ?뱀씤 ?댁뿭
            day_column = 0
            value_column = 6
            vendor_type = "2"
            time_format = "%Y%m%d%H%M%S"
            total_day_val = str(ws["A2"].value)
            
            try:
                fn_tick = datetime.datetime.strptime(total_day_val, time_format)
            except ValueError:
                try:
                    fn_tick = datetime.datetime.strptime(total_day_val, "%Y/%m/%d  %H:%M")
                    time_format = "%Y/%m/%d  %H:%M"
                except ValueError:
                    try:
                        fn_tick = datetime.datetime.strptime(total_day_val, "%Y%m%d")
                        time_format = "%Y%m%d"
                    except ValueError:
                        logs.warning(f"Shinhan date parse failed on {total_day_val}")
                        fn_tick = datetime.datetime.now()
            
            temp_month = fn_tick.month
            # Advance to first day of next month
            while temp_month == fn_tick.month:
                fn_tick += datetime.timedelta(days=1)
            fn_tick = fn_tick.replace(day=14, hour=17, minute=42, second=0)

    elif vendor == "samsung":
        day_column = 2
        value_column = 5
        vendor_type = "2"
        time_format = "%Y.%m.%d"
        try:
            val = str(ws["L2"].value).strip()
            if val == "":
                datetime.datetime.now().replace(year=year, month=int(sheetname[-2:]), day=13, hour=18, minute=3, second=0) 
            else: fn_tick = datetime.datetime.strptime(val, "%Y%m%d").replace(hour=18, minute=3, second=0)
        except Exception as e:
            logs.warning(f"Samsung total_day parse fail: {e}")

    elif vendor == "KB":
        if str(ws["T7"].value).strip() == "?뱀씤踰덊샇":
            total_col = "S"
            value_column = 7
            vendor_type = "4"
            temp_day = 0
        elif str(ws["F9"].value).strip() == "?댁슜湲덉븸":
            total_col = "M"
            value_column = 5
            vendor_type = "6"
            temp_day = 10
        else:
            total_col = "M"
            value_column = 5
            vendor_type = "5"
            temp_day = 1
            
        day_column = 0
        time_format = "%Y-%m-%d"
        max_search = 100
        while max_search > 0:
            val = ws[f"{total_col}{temp_day}"].value
            try:
                if val:
                    fn_tick = datetime.datetime.strptime(str(val), time_format).replace(hour=18, minute=4, second=0)
                    break
            except ValueError:
                pass
            temp_day += 1
            max_search -= 1

    elif vendor == "woori":
        total_col = "S"
        day_column = 0
        value_column = 16
        vendor_type = "3"
        time_format = "%Y.%m.%d %H:%M:%S"
        temp_day = 1
        
        while True:
            val = ws[f"{total_col}{temp_day}"].value
            if val is not None:
                try:
                    fn_tick = datetime.datetime.strptime(str(val), "%Y.%m.%d").replace(hour=18, minute=4, second=1)
                    break
                except ValueError:
                    pass
                    
            temp_day += 1
            if temp_day > 10:
                if total_col == "J":
                    logs.error(f"Problem while reading woori sheet {sheetname}")
                    return "ERROR", Exception("Woori sheet format unrecognizable")
                total_col = "J"
                temp_day = 1
                value_column = 7

    elif vendor == "hana":
        try:
            month = int(sheetname[-2:])
        except ValueError:
            month = 1
            
        the_day = datetime.date(year=year, month=month, day=28) + datetime.timedelta(weeks=1)
        the_day = the_day.replace(day=13)
        
        day_column = 0
        value_column = 5
        vendor_type = "4"
        time_format = "%Y.%m.%d"
        fn_tick = datetime.datetime.combine(the_day, datetime.time(18, 3, 0))

    return Analysis_sheet(ws, vendor, year, day_column, value_column, vendor_type, time_format, fn_tick)


def duplicateDates(datas: List[list]) -> List[list]:
    """Aggregate adjacent transactions with the exact same datetime.
    
    Operates in O(N) rather than the original O(N^2) list.remove logic.
    """
    print("\nProcessing same date", end="", flush=True)
    if not datas:
        return datas

    datas.sort(key=lambda x: x[0])
    
    result = [datas[0]]
    for item in datas[1:]:
        if result[-1][0] == item[0]:
            # Same timestamp -> Aggregate values
            result[-1][1] += item[1]
        else:
            result.append(item)
            
    print(" ... duplicated datas aggregated!")
    return result


def _correct_pickle_years(key: str, data: List[list]) -> List[list]:
    """Pre-2000 ?좎쭨瑜??쒗듃紐낆뿉????궛???곕룄濡??뺤젙?쒕떎.

    ?덉떆: '?곕━2312' ??year=2023, '?곕━2401' ??year=2024
    """
    try:
        correct_year = 2000 + int(key[-4:-2])
    except (ValueError, IndexError):
        return data  # ?곕룄瑜??????놁쑝硫?洹몃?濡?

    result = []
    for entry in data:
        dt, val = entry[0], entry[1]
        if dt.year < 2000:
            try:
                fixed_dt = dt.replace(year=correct_year)
                logs.warning(f"pickle '{key}': ?좎쭨 蹂댁젙 {dt.strftime('%m-%d %H:%M:%S')} ??{fixed_dt}")
                result.append([fixed_dt, val])
            except ValueError:
                result.append(entry)  # 蹂댁젙 遺덇? ???먮낯 ?좎?
        else:
            result.append(entry)
    return result


def main(cur: Any, filepath_exel: str = "移대뱶?듯빀.xlsx") -> None:
    """Main card integration pipeline."""
    datas: List[list] = []

    if os.path.isfile("cards.pickle"):
        with open("cards.pickle", "rb") as f:
            pickleData = pickle.load(f)
    else:
        pickleData = {}

    logs.msg("Opening the excel file")
    try:
        wb = openpyxl.load_workbook(filepath_exel)
    except Exception as e:
        logs.error(f"Excel 濡쒕뱶 ?ㅽ뙣 ({filepath_exel}): {e}")
        raise

    logs.msg("Processing sheets")
    
    # Check old pickle data for sheets that no longer exist in the Excel file
    for key, cached_data in pickleData.items():
        if key not in wb.sheetnames:
            # 1900?꾨? ?좎쭨瑜??쒗듃紐?湲곕컲?쇰줈 ?곕룄 蹂댁젙
            corrected = _correct_pickle_years(key, cached_data)
            datas.extend(corrected)

    delSheetList = []
    
    for sheetname in wb.sheetnames:
        if sheetname == "Info":
            continue
            
        ws = wb[sheetname]
        data = card_analysis(sheetname, ws)

        if isinstance(data, tuple) and data[0] == "ERROR":
            logs.warning(f"Sheet {sheetname} ?ㅽ궢?? {data[1]}")
            continue

        # 1900?꾨? ?좎쭨瑜??쒗듃紐낆쑝濡?蹂댁젙 (?좉퇋 遺꾩꽍 ?곗씠?곕룄 defensive 泥댄겕)
        clean_data = _correct_pickle_years(sheetname, list(data))  # type: ignore

        pickleData[sheetname] = copy.copy(clean_data)
        datas.extend(clean_data)

        # Parse year/month carefully
        try:
            mm_str = sheetname[-2:]
            yy_str = sheetname[-4:-2]
            mm = int(mm_str)
            yy = 2000 + int(yy_str)
            
            # If month < 11, it implies it's +2 month delta, if >=11 it wraps
            # Keep original logic:
            if mm < 11:
                calc_mm = mm + 2
                calc_yy = yy
            else:
                calc_mm = mm - 10
                calc_yy = yy + 1
                
            expire_date = datetime.date(calc_yy, calc_mm, 1) - datetime.timedelta(days=1)
            
            if expire_date < datetime.date.today():
                delSheetList.append(sheetname)
                
        except ValueError:
            pass # Keep if filename parse fails
            
    # pickle ?????湲곗〈 罹먯떆??蹂댁젙 (?ㅼ쓬 ?ㅽ뻾遺???щ컮瑜??곗씠???ъ슜)
    for key in list(pickleData.keys()):
        pickleData[key] = _correct_pickle_years(key, pickleData[key])

    # Save back partial state
    with open("cards.pickle", "wb") as f:
        pickle.dump(pickleData, f)

    datas = duplicateDates(datas)


    totals = sum(d[1] for d in datas)
    print(f"\n珥?湲덉븸 ?⑷퀎 泥댄겕: {totals}")
    if totals != 0:
        err_msg = f"[aborted] Calculation mismatch. Please check again... {totals}"
        logs.error(err_msg)
        raise ValueError("totals have to be 0")

    # DB Write
    try:
        cur.execute("SELECT date FROM accounts_cards LIMIT 1")
        cur.execute("DELETE FROM accounts_cards")
    except psycopg2.OperationalError:
        cur.execute(
            "CREATE TABLE accounts_cards (date TEXT, balance INTEGER, PRIMARY KEY(date))"
        )

    # Use normalized portfolio history when available and fall back to the
    # legacy summary table during the transition.
    min_max = None
    try:
        cur.execute(
            "SELECT min(recorded_at), max(recorded_at) FROM portfolio_balance_history"
        )
        min_max = cur.fetchone()
    except Exception:
        min_max = None

    if min_max and min_max[0]:
        start = min_max[0]
        end = min_max[1]
    else:
        try:
            cur.execute("SELECT min(date), max(date) FROM accounts_balance")
            min_max = cur.fetchone()
            if min_max and min_max[0]:
                start = datetime.datetime.strptime(min_max[0], "%Y-%m-%d %H:%M:%S")
                end = datetime.datetime.strptime(min_max[1], "%Y-%m-%d %H:%M:%S")
            else:
                logs.warning("No balance history found; skipping accounts_cards insert.")
                return
        except Exception:
            logs.warning("No balance history table found; skipping accounts_cards insert.")
            return

    insert_rows = []
    prev = 0
    for n, row_data in enumerate(datas):
        tick, amt = row_data[0], row_data[1]
        prev += amt

        if tick < start:
            continue
        if tick > end:
            break

        insert_rows.append((tick.strftime("%Y-%m-%d %H:%M:%S"), prev))
        if n % 50 == 0:
            print(".", end="", flush=True)

    if insert_rows:
        print("[Start] Executing DB", flush=True)
        psycopg2.extras.execute_values(
            cur,
            "INSERT INTO accounts_cards (date, balance) VALUES %s ON CONFLICT (date) DO NOTHING",
            insert_rows
        )
    print(f"\n[accounts_cards] {len(insert_rows)}嫄??쎌엯 ?꾨즺", flush=True)

    # Auto-cleanup old sheets
    if delSheetList:
        logs.msg(f"\nDeleting expired sheets: {delSheetList}")
        for ds in delSheetList:
            del wb[ds]
        wb.save(filename=filepath_exel)
        
    logs.msg("\nDone")


if __name__ == "__main__":
    con = psycopg2.connect(
        host=os.getenv("DB_HOST", "localhost"),
        database=os.getenv("DB_NAME", "total_account"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        port=os.getenv("DB_PORT", "5432")
    )
    con.set_session(autocommit=True)
    cursor = con.cursor()
    try:
        main(cursor)
        # Uncomment to commit manually when running standalone
        # con.commit()
    except Exception as e:
        logs.error(f"?ㅽ뻾 以?援ъ“???먮윭 諛쒖깮: {e}")
    finally:
        con.close()
